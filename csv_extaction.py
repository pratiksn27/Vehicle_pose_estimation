import os
from openpyxl import Workbook
import openpyxl

data = os.listdir(r"C:\Users\prati\PycharmProjects\AIVEHICLEPOSE\AI my priject\90")
data3=os.listdir(r"C:\Users\prati\PycharmProjects\AIVEHICLEPOSE\AI my priject\270")


print(len(data3))

#book=Workbook()

book2=openpyxl.load_workbook(r'C:\Users\prati\PycharmProjects\AIVEHICLEPOSE\Data.xlsx')
sheet=book2.active
for i in range(0,269):
    sheet.cell(row=i+688, column=1).value=data3[i]


book2.save(r'C:\Users\prati\PycharmProjects\AIVEHICLEPOSE\Data.xlsx')



#a3=sheet['A3']
#print(a3.value)


